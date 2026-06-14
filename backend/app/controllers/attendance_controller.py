from fastapi import UploadFile
from bson import ObjectId

from datetime import (
    datetime,
    timezone,
    timedelta
)

import os
import numpy as np

from app.config.db import (

    attendance_collection,

    registered_students_collection,

    classes_collection,
    
    students_collection,
    
    attendance_sessions_collection
)


from app.utils.distance import (
    calculate_distance
)

from app.utils.face_recognition_utils import (
    get_face_encoding
)

from app.controllers.session_controller import (
    verify_attendance_session
)

from app.utils.sheet_export import (
    append_attendance_sheet,
    sync_csv_manual_override
)

# =========================
# 📍 COLLEGE LOCATION
# =========================
# COLLEGE_LAT = 22.576028
# COLLEGE_LNG = 88.427458

# COLLEGE_LAT = 22.601870
# COLLEGE_LNG = 88.435425

#Home
COLLEGE_LAT = 23.526444
COLLEGE_LNG = 87.742552
# =========================
# 🔥 FACE MATCH THRESHOLD
# =========================
FACE_THRESHOLD = 0.5


# =========================
# ✅ MARK ATTENDANCE
# =========================
def mark_attendance(

    user_id,
    lat,
    lng,
    session_code,
    session_uuid,
    file: UploadFile,
    classroom_beacon: str = None,
    otp_code: str = None
):

    # =========================
    # 👨‍🎓 GET STUDENT
    # =========================
    student = (
        registered_students_collection
        .find_one({

            "_id":
                ObjectId(user_id)
        })
    )

    if not student:

        return {

            "status": "Error",

            "error":
                "Student not found"
        }

    # =========================
    # 📍 LOCATION CHECK
    # =========================
    distance = calculate_distance(

        lat,
        lng,

        COLLEGE_LAT,
        COLLEGE_LNG
    )

    if distance > 0.2:

        return {

            "status": "Error",

            "error":
                "Outside college area"
        }

    # =========================
    # 📡 VERIFY ATTENDANCE SESSION & GET SESSION INFO
    # =========================
    session_info = verify_attendance_session(
        session_code,
        session_uuid,
        classroom_beacon,
        otp_code
    )

    if not session_info:

        return {

            "status": "Error",

            "error":
                "No active session, classroom beacon mismatch, or incorrect manual verification code"
        }
        
    session_class_id = session_info["class_id"]

    # =========================
    # 📂 SAVE IMAGE
    # =========================
    os.makedirs(
        "uploads",
        exist_ok=True
    )

    file_path = (

        f"uploads/"
        f"{student['roll']}_attendance.jpg"
    )

    with open(file_path, "wb") as f:

        f.write(file.file.read())

    # =========================
    # 🤖 GET FACE ENCODING
    # =========================
    encoding = get_face_encoding(
        file_path
    )

    if encoding is None:

        return {

            "status": "Error",

            "error":
                "No face detected"
        }

    # =========================
    # 🔍 MATCH FACE
    # =========================
    stored_encoding = np.array(

        student["face_encoding"]
    )

    distance_score = np.linalg.norm(

        stored_encoding -

        np.array(encoding)
    )

    print(
        "🔍 FACE DISTANCE:",
        distance_score
    )

    if distance_score > FACE_THRESHOLD:

        return {

            "status": "Error",

            "error":
                "Face does not match"
        }

    # =========================
    # 📅 TODAY DATE (IST)
    # =========================
    ist_time = datetime.now(timezone.utc) + timedelta(hours=5, minutes=30)
    today = ist_time.strftime("%Y-%m-%d")

    # =========================
    # ❌ ALREADY PRESENT IN LAST 5 MINUTES?
    # =========================
    five_minutes_ago = datetime.now(timezone.utc) - timedelta(minutes=5)
    existing = (
        attendance_collection.find_one({
            "student_id": user_id,
            "status": "Present",
            "created_at": {
                "$gte": five_minutes_ago
            }
        })
    )

    if existing:

        return {

            "status": "Error",

            "error":
                "Attendance already marked in the last 5 minutes"
        }

    # =========================
    # 🏫 FIND STUDENT CLASS
    # =========================
    class_data = classes_collection.find_one({
        "_id": ObjectId(session_class_id)
    })

    enrolled = students_collection.find_one({
        "class_id": session_class_id,
        "roll": student["roll"]
    })

    if not class_data or not enrolled:
        return {
            "status": "Error",
            "error": "You are not enrolled in the class for this session"
        }

    class_id = str(class_data["_id"])
    class_name = class_data["class_name"]

    # =========================
    # ✅ SAVE ATTENDANCE
    # =========================
    now_ist = datetime.now(
        timezone(timedelta(hours=5, minutes=30))
    )

    session_created_at = session_info.get("created_at")
    if session_created_at:
        ist_timezone = timezone(timedelta(hours=5, minutes=30))
        if session_created_at.tzinfo is None:
            session_created_at = session_created_at.replace(tzinfo=timezone.utc)
        session_ist = session_created_at.astimezone(ist_timezone)
        session_time_str = session_ist.strftime("%Y-%m-%d %H:%M:%S")
    else:
        session_time_str = today + " " + now_ist.strftime("%H:%M:%S")

    attendance_collection.update_one(
        {
            "class_id": class_id,
            "roll": student["roll"],
            "date": today
        },
        {
            "$set": {
                "student_id": user_id,
                "name": student["name"],
                "roll": student["roll"],
                "class_id": class_id,
                "class_name": class_name,
                "date": today,
                "session_uuid": session_uuid,
                "session_time": session_time_str,
                "status": "Present"
            },
            "$setOnInsert": {
                "created_at": datetime.now(timezone.utc)
            }
        },
        upsert=True
    )

    # =========================
    # 📝 APPEND TO CSV
    # =========================
    department = class_data.get("department", "") if class_data else ""
    append_attendance_sheet(
        student_id=user_id,
        name=student["name"],
        roll=student["roll"],
        class_name=class_name,
        date=today,
        time=now_ist.strftime("%I:%M %p"),
        section=class_data.get("section", "") if class_data else "",
        department=department,
        course_code=class_data.get("course_code", "N/A") if class_data else "N/A",
        semester=str(class_data.get("semester", "N/A")) if class_data else "N/A",
        year=str(class_data.get("year", "N/A")) if class_data else "N/A",
        academic_session=class_data.get("academic_session", "N/A") if class_data else "N/A",
    )

    # =========================
    # 🔄 UPDATE REALTIME STUDENTS COLLECTION
    # =========================
    students_collection.update_one(
        {"roll": student["roll"], "class_id": class_id},
        {"$set": {"attendance_status": "Present"}}
    )

    # =========================
    # 🔄 UPDATE LAST ATTENDANCE
    # =========================
    registered_students_collection.update_one(

        {
            "_id":
                ObjectId(user_id)
        },

        {
            "$set": {

                "last_attendance":
                    today
            }
        }
    )

    return {

        "status": "Success",

        "message":
            "Attendance marked successfully"
    }


# =========================
# ❌ UNMARK ATTENDANCE
# =========================
def unmark_attendance(user_id):
    ist_time = datetime.now(timezone.utc) + timedelta(hours=5, minutes=30)
    today = ist_time.strftime("%Y-%m-%d")

    # Get student to find enrolled classes
    from bson import ObjectId
    student_doc = registered_students_collection.find_one({"_id": ObjectId(user_id)})
    roll = student_doc.get("roll") if student_doc else None

    # Find active session or most recent session of today
    session_query = {"active": True}
    if roll:
        # Get classes student is enrolled in to narrow down session
        enrolled_classes = list(students_collection.find({"roll": roll}))
        class_ids = [str(c["class_id"]) for c in enrolled_classes]
        if class_ids:
            session_query["class_id"] = {"$in": class_ids}

    active_session = attendance_sessions_collection.find_one(session_query)
    
    if active_session:
        session_uuid = active_session.get("session_uuid")
        query = {"student_id": user_id, "session_uuid": session_uuid}
    else:
        # Fallback to the latest record of today
        query = {"student_id": user_id, "date": today}

    record = attendance_collection.find_one(query)
    result = attendance_collection.delete_one(query)

    if result.deleted_count > 0:
        if record:
            class_id = record.get("class_id")
            class_doc = classes_collection.find_one({"_id": ObjectId(class_id)}) if class_id else None
            if class_doc:
                sync_csv_manual_override(
                    class_name=class_doc.get("class_name", ""),
                    roll=record.get("roll", ""),
                    date=record.get("date", today),
                    new_status="Absent",
                    department=class_doc.get("department", "")
                )

        return {
            "status": "Success",
            "message": "Attendance Unmarked"
        }
    else:
        return {
            "status": "Error",
            "error": "Attendance not found"
        }
# =========================
# 📊 TODAY STATUS
# =========================
def get_today_status(user_id):
    latest_attendance = (
        attendance_collection.find_one(
            {"student_id": user_id},
            sort=[("created_at", -1)]
        )
    )

    remaining_minutes = 0
    remaining_seconds = 0
    if latest_attendance and "created_at" in latest_attendance and latest_attendance.get("status") == "Present":
        created_at = latest_attendance["created_at"]
        if created_at.tzinfo is None:
            created_at = created_at.replace(tzinfo=timezone.utc)
        
        now = datetime.now(timezone.utc)
        elapsed = (now - created_at).total_seconds()
        
        if elapsed < 300: # 5 minutes cooldown
            remaining_seconds = max(0, int(300 - elapsed))
            remaining_minutes = int(remaining_seconds // 60) + (1 if remaining_seconds % 60 > 0 else 0)

    if remaining_seconds > 0:
        actual_status = latest_attendance.get("status", "Present") if latest_attendance else "Present"
        return {
            "marked": True,
            "status": actual_status,
            "remaining_minutes": remaining_minutes,
            "remaining_seconds": remaining_seconds
        }

    return {
        "marked": False,
        "status": "Absent",
        "remaining_minutes": 0,
        "remaining_seconds": 0
    }


# =========================
# 📜 WEEKLY HISTORY
# =========================
def get_weekly_history(user_id):
    from app.config.db import registered_students_collection
    from bson import ObjectId

    student = registered_students_collection.find_one({"_id": ObjectId(user_id)})
    roll = student.get("roll") if student else None

    query = {"student_id": user_id}
    if roll:
        query = {
            "$or": [
                {"student_id": user_id},
                {"roll": roll}
            ]
        }

    history = list(
        attendance_collection.find(query)
        .sort(
            "created_at",
            -1
        )
    )

    formatted = []

    for item in history:
        created_at_str = None
        if "created_at" in item:
            created_at_str = item["created_at"].isoformat()

        formatted.append({
            "date": item["date"],
            "status": item["status"],
            "time": created_at_str
        })

    return {
        "history": formatted
    }


# =========================
# 📊 CLASS ATTENDANCE REPORT
# =========================
def get_class_attendance_report(
    class_id,
    report_date=None
):

    # =========================
    # 🏫 FIND CLASS
    # =========================
    class_data = classes_collection.find_one({

        "_id":
            ObjectId(class_id)
    })

    if not class_data:

        return {

            "success": False,

            "error":
                "Class not found"
        }

    students = class_data.get(
        "students",
        []
    )

    # =========================
    # 📅 AVAILABLE DATES
    # =========================
    distinct_sessions = attendance_collection.distinct("session_time", {"class_id": class_id})
    distinct_sessions = [s for s in distinct_sessions if s]
    
    distinct_dates = attendance_collection.distinct("date", {"class_id": class_id})
    
    all_dates_set = set(distinct_sessions)
    for d in distinct_dates:
        if not any(s.startswith(d) for s in distinct_sessions):
            all_dates_set.add(d)
            
    all_dates = sorted(list(all_dates_set), reverse=True)

    # =========================
    # 📅 DETERMINE TARGET DATE
    # =========================
    ist_now = datetime.now(timezone.utc) + timedelta(hours=5, minutes=30)
    today_str = ist_now.strftime("%Y-%m-%d")
    
    if not report_date:
        if all_dates:
            target_date = all_dates[0]
        else:
            target_date = today_str
    else:
        target_date = report_date

    report = []

    present_count = 0
    absent_count = 0
    na_count = 0

    # =========================
    # 👨‍🎓 LOOP STUDENTS
    # =========================
    for student in students:

        roll = student.get("roll")

        name = student.get("name")

        registered_student = (
            registered_students_collection.find_one({
                "roll": roll
            })
        )

        # =========================
        # ❌ NOT REGISTERED
        # =========================
        if not registered_student:
            status = "N/A"
            total_count = 0
            time_str = "N/A"
            na_count += 1
        else:
            # =========================
            # 📅 DATE-SPECIFIC ATTENDANCE
            # =========================
            if " " in target_date:
                attendance = (
                    attendance_collection.find_one({
                        "class_id": class_id,
                        "roll": roll,
                        "session_time": target_date
                    })
                )
            else:
                attendance = (
                    attendance_collection.find_one({
                        "class_id": class_id,
                        "roll": roll,
                        "date": target_date
                    })
                )

            time_str = "N/A"

            if attendance and attendance.get("status") == "Present":
                status = "Present"
                present_count += 1
                created_at = attendance.get("created_at")
                if created_at:
                    ist_timezone = timezone(timedelta(hours=5, minutes=30))
                    if created_at.tzinfo is None:
                        created_at = created_at.replace(tzinfo=timezone.utc)
                    ist_time = created_at.astimezone(ist_timezone)
                    time_str = ist_time.strftime("%H:%M:%S")
            else:
                status = "Absent"
                absent_count += 1

        # =========================
        # 📊 TOTAL COUNT
        # =========================
        total_count = attendance_collection.count_documents({
            "class_id": class_id,
            "roll": roll,
            "status": "Present"
        })

        report.append({
            "name": name,
            "roll": roll,
            "attendance_status": status,
            "time": time_str,
            "total_attendance": total_count
        })

    # =========================
    # 📊 TOTAL CLASSES HELD
    # =========================
    total_classes = len(all_dates)

    return {

        "success": True,

        "total_monthly_classes": total_classes,

        "class_name":
            class_data["class_name"],

        "department":
            class_data.get("department", ""),

        "section":
            class_data.get("section", ""),

        "academic_session":
            class_data.get("academic_session", ""),

        "course_code":
            class_data.get("course_code", ""),

        "present_students":
            present_count,

        "absent_students":
            absent_count,

        "na_students":
            na_count,

        "report_date":
            target_date,

        "available_dates":
            all_dates,

        "students":
            report
    }

# =========================
# 🎓 GET ATTENDED CLASSES
# =========================
def get_student_classes(user_id):
    pipeline = [
        {"$match": {"student_id": user_id, "status": "Present"}},
        {"$group": {"_id": "$class_name", "attended_count": {"$sum": 1}}},
        {"$sort": {"attended_count": -1}}
    ]
    results = list(attendance_collection.aggregate(pipeline))
    
    classes_list = []
    for item in results:
        # Use class_name string
        c_name = item["_id"] if item["_id"] else "Unknown Class"
        classes_list.append({
            "class_name": c_name,
            "attended_count": item["attended_count"]
        })
        
    return {
        "success": True,
        "classes": classes_list
    }



# =========================
# ✍ UPDATE MANUAL ATTENDANCE
# =========================
def update_manual_attendance(class_id: str, students: list):
    # Get current IST date
    now = datetime.now(timezone.utc)
    ist_time = now + timedelta(hours=5, minutes=30)
    today_str = ist_time.strftime("%Y-%m-%d")

    class_doc = classes_collection.find_one({"_id": ObjectId(class_id)})
    class_name = class_doc.get("class_name", "Unknown Class") if class_doc else "Unknown Class"

    # Find active session or most recent session of today/overall
    session = attendance_sessions_collection.find_one({
        "class_id": class_id,
        "active": True
    })
    if not session:
        session = attendance_sessions_collection.find_one(
            {"class_id": class_id},
            sort=[("created_at", -1)]
        )
        
    session_uuid = session.get("session_uuid") if session else None
    session_time_str = None
    if session:
        session_created_at = session.get("created_at")
        if session_created_at:
            ist_timezone = timezone(timedelta(hours=5, minutes=30))
            if session_created_at.tzinfo is None:
                session_created_at = session_created_at.replace(tzinfo=timezone.utc)
            session_ist = session_created_at.astimezone(ist_timezone)
            session_time_str = session_ist.strftime("%Y-%m-%d %H:%M:%S")
            
    if not session_time_str:
        session_time_str = today_str + " 00:00:00"

    for student in students:
        roll = student.roll
        status = student.attendance_status
        name = student.name

        # Find student to get student_id if possible
        s_doc = registered_students_collection.find_one({"roll": roll})
        student_id = str(s_doc["_id"]) if s_doc else None

        # Enforce one record per day per class
        query = {
            "class_id": class_id,
            "roll": roll,
            "date": today_str
        }

        update_data = {
            "status": status,
            "name": name,
            "roll": roll,
            "class_id": class_id,
            "class_name": class_name,
            "date": today_str,
            "session_time": session_time_str
        }
        if session_uuid:
            update_data["session_uuid"] = session_uuid
        
        if student_id:
            update_data["student_id"] = student_id

        attendance_collection.update_one(
            query,
            {
                "$set": update_data,
                "$setOnInsert": {"created_at": now}
            },
            upsert=True
        )

        # Update the realtime students collection so Admin UI reflects it
        students_collection.update_one(
            {"roll": roll, "class_id": class_id},
            {"$set": {"attendance_status": status}}
        )

        department = class_doc.get("department", "") if class_doc else ""
        sync_csv_manual_override(
            class_name=class_name,
            roll=roll,
            date=today_str,
            new_status=status,
            time=now.astimezone(ist_timezone).strftime("%I:%M %p"),
            department=department,
            section=class_doc.get("section", "") if class_doc else "",
            course_code=class_doc.get("course_code", "N/A") if class_doc else "N/A",
            semester=str(class_doc.get("semester", "N/A")) if class_doc else "N/A",
            year=str(class_doc.get("year", "N/A")) if class_doc else "N/A",
            academic_session=class_doc.get("academic_session", "N/A") if class_doc else "N/A",
        )

    return {"message": "Attendance updated successfully"}

# =========================
# 📥 EXPORT CUSTOM EXCEL REPORT
# =========================
def export_custom_attendance_excel(class_id: str):
    from fastapi.responses import StreamingResponse
    import io
    import re
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    class_doc = classes_collection.find_one({"_id": ObjectId(class_id)})
    if not class_doc:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Class not found")
        
    class_name = class_doc.get("class_name", "Class")
    department = class_doc.get("department", "Dept")
    course_code = class_doc.get("course_code", "N/A")
    semester = class_doc.get("semester", "N/A")
    section = class_doc.get("section", "N/A")
    academic_session = class_doc.get("academic_session", "N/A")
    
    # Determine academic session start date (June of start year)
    start_year = None
    if academic_session:
        # Match a 4-digit year
        match = re.search(r'\b(19|20)\d{2}\b', str(academic_session))
        if match:
            start_year = match.group(0)

    session_start_date = None
    if start_year:
        session_start_date = f"{start_year}-06-01"

    # Build query
    query = {"class_id": class_id}
    if session_start_date:
        query["date"] = {"$gte": session_start_date}
        
    # Get all distinct sessions (and legacy dates) for this class in this range
    distinct_sessions = attendance_collection.distinct("session_time", query)
    distinct_sessions = [s for s in distinct_sessions if s]
    
    distinct_dates = attendance_collection.distinct("date", query)
    
    all_columns_set = set(distinct_sessions)
    for d in distinct_dates:
        if not any(s.startswith(d) for s in distinct_sessions):
            all_columns_set.add(d)
            
    columns = sorted(list(all_columns_set))
    
    # Initialize students_dict with ALL students currently in the class
    students_dict = {}
    class_students = list(students_collection.find({"class_id": class_id}))
    for s in class_students:
        roll = s.get("roll")
        students_dict[roll] = {
            "name": s.get("name", ""),
            "section": section,
            "dates": {},
            "presents": 0,
            "absents": 0
        }
    
    # Get all records
    records = list(attendance_collection.find(query))
    
    # Populate attendance data
    for r in records:
        roll = r.get("roll")
        if roll not in students_dict:
            students_dict[roll] = {
                "name": r.get("name", ""),
                "section": section,
                "dates": {},
                "presents": 0,
                "absents": 0
            }
        
        status = r.get("status", "Absent")
        session_key = r.get("session_time") or r.get("date")
        time_str = r.get("time", "")
        students_dict[roll]["dates"][session_key] = {
            "status": status,
            "time": time_str
        }
        if status == "Present":
            students_dict[roll]["presents"] += 1
        elif status == "Absent":
            students_dict[roll]["absents"] += 1

    # Generate Excel Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Register"
    ws.views.sheetView[0].showGridLines = True
    
    # Fonts
    title_font = Font(name="Segoe UI", size=20, bold=True, color="FFFFFF")
    meta_label_font = Font(name="Segoe UI", size=11, bold=True, color="6B7280")
    meta_value_font = Font(name="Segoe UI", size=11, bold=True, color="1F2937")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=11, color="374151")
    data_bold_font = Font(name="Segoe UI", size=11, bold=True, color="111827")
    
    # Alignments
    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    title_align = Alignment(horizontal="center", vertical="center")
    
    # Fills
    header_fill = PatternFill(start_color="2E2B42", end_color="2E2B42", fill_type="solid")
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    present_fill = PatternFill(start_color="E6F9F0", end_color="E6F9F0", fill_type="solid")
    absent_fill = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
    summary_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    
    title_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo 600

    # Borders
    thin_border_side = Side(style='thin', color='E5E7EB')
    thick_bottom_side = Side(style='medium', color='4F46E5')
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    header_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thick_bottom_side)
    
    # 1. Title
    ws.merge_cells("A1:G2")
    title_cell = ws["A1"]
    title_cell.value = f"ATTENDANCE REGISTER: {class_name.upper()}"
    title_cell.font = title_font
    title_cell.alignment = title_align
    title_cell.fill = title_fill
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 25
    
    # 2. Metadata Block
    metadata = [
        [("Course Code:", course_code), ("Department:", department), ("Semester:", semester)],
        [("Section:", section), ("Academic Session:", academic_session), ("Total Students:", len(students_dict))]
    ]
    
    for r_idx, row_data in enumerate(metadata, start=3):
        ws.row_dimensions[r_idx].height = 20
        c_idx = 1
        for label, val in row_data:
            # Label
            ws.cell(row=r_idx, column=c_idx, value=label).font = meta_label_font
            ws.cell(row=r_idx, column=c_idx).alignment = left_align
            # Value
            ws.cell(row=r_idx, column=c_idx+1, value=val).font = meta_value_font
            ws.cell(row=r_idx, column=c_idx+1).alignment = left_align
            c_idx += 2
            
    # Empty Row
    ws.row_dimensions[5].height = 10
    
    # Format Dates for Headers
    def format_date_header(date_str):
        try:
            if "T" in date_str or len(date_str) > 10:
                dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                return dt.strftime("%d %b")
            else:
                dt = datetime.strptime(date_str, "%Y-%m-%d")
                return dt.strftime("%d %b")
        except:
            return str(date_str)[:10]

    formatted_columns = [format_date_header(c) for c in columns]

    # 3. Table Headers
    headers = ["Student Name", "Roll Number", "Section"] + formatted_columns + ["Total Presents", "Total Absents", "Percentage"]
    ws.row_dimensions[6].height = 30
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=c_idx, value=h)
        cell.font = header_font
        
        # Highlight summary column headers with a distinct color
        if c_idx > len(columns) + 3:
            cell.fill = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
        else:
            cell.fill = header_fill
            
        cell.alignment = center_align if c_idx > 1 else left_align
        cell.border = header_border
        
    # FREEZE PANES: Locks rows 1-6 only
    ws.freeze_panes = "A7"
        
    # 4. Data Rows
    def roll_sort_key(roll_str):
        nums = re.findall(r'\d+', str(roll_str))
        return int(nums[0]) if nums else str(roll_str)
        
    row_idx = 7
    for roll in sorted(students_dict.keys(), key=roll_sort_key):
        ws.row_dimensions[row_idx].height = 25
        is_even = (row_idx % 2 == 0)
        row_fill = zebra_fill if is_even else None
        
        # Student Name
        c_name = ws.cell(row=row_idx, column=1, value=students_dict[roll]["name"])
        c_name.font = data_font
        c_name.alignment = left_align
        c_name.border = cell_border
        if row_fill: c_name.fill = row_fill
        
        # Roll Number
        c_roll = ws.cell(row=row_idx, column=2, value=roll)
        c_roll.font = data_font
        c_roll.alignment = center_align
        c_roll.border = cell_border
        if row_fill: c_roll.fill = row_fill
        
        # Section
        c_sec = ws.cell(row=row_idx, column=3, value=students_dict[roll]["section"])
        c_sec.font = data_font
        c_sec.alignment = center_align
        c_sec.border = cell_border
        if row_fill: c_sec.fill = row_fill
        
        # Attendance Status columns
        col_idx = 4
        for col in columns:
            date_info = students_dict[roll]["dates"].get(col, {})
            
            if isinstance(date_info, dict):
                status = date_info.get("status", "-")
                time_str = date_info.get("time", "")
            else:
                status = date_info
                time_str = ""
            
            if status == "Present":
                if time_str:
                    try:
                        from datetime import datetime
                        time_obj = datetime.strptime(time_str, "%H:%M:%S")
                        formatted_time = time_obj.strftime("%I:%M %p")
                        val = f"P\n({formatted_time})"
                    except:
                        val = f"P\n({time_str})"
                else:
                    val = "P"
            elif status == "Absent":
                if time_str:
                    try:
                        from datetime import datetime
                        time_obj = datetime.strptime(time_str, "%H:%M:%S")
                        formatted_time = time_obj.strftime("%I:%M %p")
                        val = f"A\n({formatted_time})"
                    except:
                        val = f"A\n({time_str})"
                else:
                    val = "A"
            else:
                val = "-"

            c_status = ws.cell(row=row_idx, column=col_idx, value=val)
            c_status.font = data_bold_font if val != "-" else data_font
            c_status.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c_status.border = cell_border
            
            # Highlight P/A
            if "P" in val:
                c_status.fill = present_fill
                c_status.font = Font(name="Segoe UI", size=10, bold=True, color="059669")
            elif "A" in val:
                c_status.fill = absent_fill
                c_status.font = Font(name="Segoe UI", size=10, bold=True, color="DC2626")
            elif row_fill:
                c_status.fill = row_fill
                
            col_idx += 1
            
        # Total Presents
        presents = students_dict[roll]["presents"]
        c_p = ws.cell(row=row_idx, column=col_idx, value=presents)
        c_p.font = data_bold_font
        c_p.alignment = center_align
        c_p.border = cell_border
        c_p.fill = summary_fill
        
        # Total Absents
        absents = students_dict[roll]["absents"]
        c_a = ws.cell(row=row_idx, column=col_idx+1, value=absents)
        c_a.font = data_bold_font
        c_a.alignment = center_align
        c_a.border = cell_border
        c_a.fill = summary_fill
        
        # Percentage
        total_sessions = presents + absents
        pct = int((presents * 100) / total_sessions) if total_sessions > 0 else 0
        c_pct = ws.cell(row=row_idx, column=col_idx+2, value=f"{pct}%")
        c_pct.font = data_bold_font
        c_pct.alignment = center_align
        c_pct.border = cell_border
        
        # Color code percentage
        if pct >= 75:
            c_pct.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
            c_pct.font = Font(name="Segoe UI", size=11, bold=True, color="065F46")
        elif pct >= 50:
            c_pct.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            c_pct.font = Font(name="Segoe UI", size=11, bold=True, color="D97706")
        else:
            c_pct.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            c_pct.font = Font(name="Segoe UI", size=11, bold=True, color="B91C1C")
            
        row_idx += 1
        
    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Determine maximum string length
        for cell in col:
            # Ignore merged cells in Title row for width calculation
            if cell.row == 1:
                continue
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
                
        # Set column width
        if c_idx == 1:
            ws.column_dimensions[col_letter].width = max(max_len + 5, 25) # Student name wider
        else:
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # Save to memory stream
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Generate filename according to user requested format
    dept_sec = f"{department}_{section}" if section and section != "N/A" else department
    session_part = academic_session if academic_session and academic_session != "N/A" else ""
    class_part = f"{class_name} ({course_code})" if course_code and course_code != "N/A" else class_name
    
    parts = [p for p in [dept_sec, session_part, class_part] if p]
    filename = ",".join(parts) + "_Register.xlsx"
    
    # Clean up any illegal Windows filename characters
    filename = re.sub(r'[\\/:*?"<>|]', "", filename).strip()
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
