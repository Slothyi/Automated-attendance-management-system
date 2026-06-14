import os
import re
import calendar
from datetime import datetime, date
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# =========================
# 📁 OUTPUT DIRECTORY
# =========================
CSV_DIR = "attendance_sheet"

# =========================
# 🎨 STYLE CONSTANTS
# =========================
TITLE_FONT        = Font(name="Segoe UI", size=15, bold=True, color="FFFFFF")
HEADER_FONT       = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
SUMMARY_FONT      = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
META_LABEL_FONT   = Font(name="Segoe UI", size=9, bold=True, color="6B7280")
META_VALUE_FONT   = Font(name="Segoe UI", size=9, bold=False, color="111827")
DATA_FONT         = Font(name="Segoe UI", size=10, color="1F2937")
DATA_BOLD_FONT    = Font(name="Segoe UI", size=10, bold=True, color="111827")
PRESENT_FONT      = Font(name="Segoe UI", size=10, bold=True, color="065F46")
ABSENT_FONT       = Font(name="Segoe UI", size=10, bold=True, color="991B1B")
DASH_FONT         = Font(name="Segoe UI", size=10, color="9CA3AF")
PCT_GREEN_FONT    = Font(name="Segoe UI", size=10, bold=True, color="065F46")
PCT_YELLOW_FONT   = Font(name="Segoe UI", size=10, bold=True, color="92400E")
PCT_RED_FONT      = Font(name="Segoe UI", size=10, bold=True, color="991B1B")

TITLE_FILL        = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
META_BG_FILL      = PatternFill(start_color="F0F0FA", end_color="F0F0FA", fill_type="solid")
META_DIVIDER_FILL = PatternFill(start_color="E0E0F5", end_color="E0E0F5", fill_type="solid")
HEADER_FILL       = PatternFill(start_color="2E2B42", end_color="2E2B42", fill_type="solid")
SUMMARY_FILL      = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
ZEBRA_FILL        = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
PRESENT_FILL      = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
ABSENT_FILL       = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
DASH_FILL         = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
PCT_GREEN_FILL    = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
PCT_YELLOW_FILL   = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
PCT_RED_FILL      = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
PRESENTS_FILL     = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
ABSENTS_FILL      = PatternFill(start_color="FFF1F2", end_color="FFF1F2", fill_type="solid")

CENTER       = Alignment(horizontal="center", vertical="center")
LEFT         = Alignment(horizontal="left",   vertical="center")
TITLE_ALIGN  = Alignment(horizontal="center", vertical="center")

THIN_SIDE      = Side(style="thin",   color="E5E7EB")
HEADER_BOTTOM  = Side(style="medium", color="4F46E5")
NONE_SIDE      = Side(style=None)
CELL_BORDER    = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
HEADER_BORDER  = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=HEADER_BOTTOM)
META_BORDER    = Border(left=NONE_SIDE, right=NONE_SIDE, top=NONE_SIDE, bottom=NONE_SIDE)


# =========================
# 🧹 SANITIZE FILENAME
# =========================
def _safe_filename(class_name: str) -> str:
    name = re.sub(r'[\\/:*?"<>|]', "", class_name)
    return name.strip() or "unknown_class"


# =========================
# 🔢 NUMERIC ROLL SORTING
# =========================
def _roll_sort_key(roll_str):
    nums = re.findall(r'\d+', str(roll_str))
    return int(nums[0]) if nums else 0


# =========================
# 📅 GENERATE DATE COLUMNS FROM JUNE 1 TO TODAY
# =========================
def _generate_date_columns(current_date_str: str, start_month: int = 6) -> list:
    current_dt = datetime.strptime(current_date_str, "%Y-%m-%d")
    year = current_dt.year
    start_year = year - 1 if current_dt.month < start_month else year
    start_date = date(start_year, start_month, 1)
    end_date = current_dt.date()

    columns = []
    d = start_date
    while d <= end_date:
        columns.append(d.strftime("%Y-%m-%d"))
        days_in_m = calendar.monthrange(d.year, d.month)[1]
        if d.day < days_in_m:
            d = date(d.year, d.month, d.day + 1)
        elif d.month == 12:
            d = date(d.year + 1, 1, 1)
        else:
            d = date(d.year, d.month + 1, 1)
    return columns


def _format_date_header(date_str: str) -> str:
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    return dt.strftime("%d %b")

def _apply_cell_style(cell, font, align, border, fill=None):
    cell.font = font
    cell.alignment = align
    cell.border = border
    if fill:
        cell.fill = fill


# =========================
# 📊 BUILD THE FORMATTED EXCEL WORKBOOK
# =========================
def _render_header_block(ws, total_cols, class_name, course_code, department, section, semester, year, academic_session, is_summary=False):
    # ROWS 1–2  ▸  TITLE BAR
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"ATTENDANCE REGISTER  ·  {class_name.upper()}"
    title_cell.font = TITLE_FONT
    title_cell.alignment = TITLE_ALIGN
    title_cell.fill = TITLE_FILL
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30
    for c in range(1, total_cols + 1):
        ws.cell(row=1, column=c).fill = TITLE_FILL
        ws.cell(row=2, column=c).fill = TITLE_FILL

    # ROW 3  ▸  THIN ACCENT LINE
    ws.row_dimensions[3].height = 4
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_cols)
    for c in range(1, total_cols + 1):
        ws.cell(row=3, column=c).fill = META_DIVIDER_FILL

    # ROWS 4–5  ▸  METADATA BLOCK
    meta_items = [
        ("Course Code",      course_code      or "N/A"),
        ("Department",       department       or "N/A"),
        ("Section",          section          or "N/A"),
        ("Semester",         semester         or "N/A"),
        ("Year",             year             or "N/A"),
        ("Academic Session", academic_session or "N/A"),
    ]

    pair_width = max(2, total_cols // 3)
    LBL_SPAN = max(1, pair_width // 2)
    VAL_SPAN = max(1, pair_width - LBL_SPAN)
    PAIR_ANCHORS = [1, 1 + pair_width, 1 + 2 * pair_width]

    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 22

    for row_offset, row_items in enumerate([meta_items[:3], meta_items[3:]]):
        excel_row = 4 + row_offset
        for c in range(1, total_cols + 1):
            ws.cell(row=excel_row, column=c).fill = META_BG_FILL

        for pair_idx, (label, value) in enumerate(row_items):
            anchor = PAIR_ANCHORS[pair_idx]
            lbl_start = anchor
            lbl_end   = anchor + LBL_SPAN - 1
            val_start = anchor + LBL_SPAN
            val_end   = total_cols if pair_idx == 2 else anchor + LBL_SPAN + VAL_SPAN - 1

            if lbl_end > lbl_start:
                try:
                    ws.merge_cells(start_row=excel_row, start_column=lbl_start, end_row=excel_row, end_column=lbl_end)
                except Exception:
                    pass
            lbl_cell = ws.cell(row=excel_row, column=lbl_start, value=f"{label.upper()} :")
            lbl_cell.font = META_LABEL_FONT
            lbl_cell.alignment = LEFT
            lbl_cell.fill = META_BG_FILL

            if val_end > val_start:
                try:
                    ws.merge_cells(start_row=excel_row, start_column=val_start, end_row=excel_row, end_column=val_end)
                except Exception:
                    pass
            val_cell = ws.cell(row=excel_row, column=val_start, value=value)
            val_cell.font = META_VALUE_FONT
            val_cell.alignment = LEFT
            val_cell.fill = META_BG_FILL

    # ROW 6  ▸  THIN ACCENT LINE
    ws.row_dimensions[6].height = 4
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=total_cols)
    for c in range(1, total_cols + 1):
        ws.cell(row=6, column=c).fill = META_DIVIDER_FILL


def _build_workbook(
    class_name: str,
    department: str,
    section: str,
    date_columns: list,
    students_data: dict,
    course_code: str = "N/A",
    semester: str = "N/A",
    year: str = "N/A",
    academic_session: str = "N/A",
):
    wb = Workbook()
    
    sorted_rolls = sorted(students_data.keys(), key=_roll_sort_key)
    
    # Pre-calculate global stats for Summary tab and Month tabs
    global_stats = {}
    for roll in sorted_rolls:
        presents = 0
        absents = 0
        for val in students_data[roll].get("dates", {}).values():
            if str(val).startswith("P"):
                presents += 1
            elif str(val).startswith("A"):
                absents += 1
        global_stats[roll] = {"presents": presents, "absents": absents}
        
    # ─────────────────────────────────────────────
    # CREATE SUMMARY SHEET
    # ─────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    summary_total_cols = 6
    _render_header_block(ws_summary, summary_total_cols, class_name, course_code, department, section, semester, year, academic_session, is_summary=True)
    
    summary_headers = ["Student Name", "Roll Number", "Section", "Total Presents (throughout Session)", "Total Absents (throughout Session)", "Percentage (throughout Session)"]
    ws_summary.row_dimensions[7].height = 32
    
    for c_idx, h in enumerate(summary_headers, start=1):
        cell = ws_summary.cell(row=7, column=c_idx, value=h)
        cell.alignment = CENTER
        cell.border = HEADER_BORDER
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row_offset, roll in enumerate(sorted_rolls):
        row_idx = 8 + row_offset
        student = students_data[roll]
        ws_summary.row_dimensions[row_idx].height = 26
        row_fill = ZEBRA_FILL if row_offset % 2 == 1 else None

        # Name, Roll, Section
        cell_name = ws_summary.cell(row=row_idx, column=1, value=student.get("name", ""))
        _apply_cell_style(cell_name, DATA_BOLD_FONT, LEFT, CELL_BORDER, row_fill)
        
        cell_roll = ws_summary.cell(row=row_idx, column=2, value=roll)
        _apply_cell_style(cell_roll, DATA_FONT, CENTER, CELL_BORDER, row_fill)
        
        cell_sec = ws_summary.cell(row=row_idx, column=3, value=student.get("section", ""))
        _apply_cell_style(cell_sec, DATA_FONT, CENTER, CELL_BORDER, row_fill)
        
        g_pres = global_stats[roll]["presents"]
        g_abs = global_stats[roll]["absents"]
        g_total = g_pres + g_abs
        g_pct = round((g_pres / g_total) * 100, 1) if g_total > 0 else 0.0
        
        cell_gp = ws_summary.cell(row=row_idx, column=4, value=g_pres)
        _apply_cell_style(cell_gp, Font(name="Segoe UI", size=10, bold=True, color="065F46"), CENTER, CELL_BORDER, PRESENTS_FILL)

        cell_ga = ws_summary.cell(row=row_idx, column=5, value=g_abs)
        _apply_cell_style(cell_ga, Font(name="Segoe UI", size=10, bold=True, color="991B1B"), CENTER, CELL_BORDER, ABSENTS_FILL)

        cell_gpct = ws_summary.cell(row=row_idx, column=6, value=f"{g_pct}%")
        _apply_cell_style(cell_gpct, PCT_GREEN_FONT if g_pct >= 75 else (PCT_YELLOW_FONT if g_pct >= 50 else PCT_RED_FONT), CENTER, CELL_BORDER, PCT_GREEN_FILL if g_pct >= 75 else (PCT_YELLOW_FILL if g_pct >= 50 else PCT_RED_FILL))

    ws_summary.column_dimensions[get_column_letter(1)].width = 25
    ws_summary.column_dimensions[get_column_letter(2)].width = 15
    ws_summary.column_dimensions[get_column_letter(3)].width = 18
    ws_summary.column_dimensions[get_column_letter(4)].width = 35
    ws_summary.column_dimensions[get_column_letter(5)].width = 35
    ws_summary.column_dimensions[get_column_letter(6)].width = 30
    
    ws_summary.freeze_panes = "A8"

    # GROUP DATES BY MONTH
    dates_by_month = defaultdict(list)
    for d in date_columns:
        dt = datetime.strptime(d, "%Y-%m-%d")
        dates_by_month[(dt.year, dt.month)].append(d)

    # ─────────────────────────────────────────────
    # CREATE MONTH SHEETS
    # ─────────────────────────────────────────────
    for (y, m), m_dates in sorted(dates_by_month.items()):
        month_name = datetime(y, m, 1).strftime("%b %Y")
        ws = wb.create_sheet(title=month_name)

        num_date_cols = len(m_dates)
        total_cols = 3 + num_date_cols + 3  # Name, Roll, Section + dates + 3 month stats
        # Removed hardcoded max(total_cols, 22) so header matches table width exactly

        _render_header_block(ws, total_cols, class_name, course_code, department, section, semester, year, academic_session, is_summary=False)

        # ROW 7  ▸  COLUMN HEADERS
        headers = ["Student Name", "Roll Number", "Section"]
        for dc in m_dates:
            headers.append(_format_date_header(dc))
        headers.extend([
            "Total Presents (in this month)", 
            "Total Absents (in this month)", 
            "Percentage (in this month)"
        ])

        ws.row_dimensions[7].height = 45 # Make header taller for long titles
        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=7, column=c_idx, value=h)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = HEADER_BORDER
            
            if c_idx > 3 + num_date_cols:
                cell.fill = SUMMARY_FILL
                cell.font = SUMMARY_FONT
            else:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT

        ws.freeze_panes = "A8"

        # ROWS 8+  ▸  DATA
        for row_offset, roll in enumerate(sorted_rolls):
            row_idx = 8 + row_offset
            student = students_data[roll]
            ws.row_dimensions[row_idx].height = 26
            row_fill = ZEBRA_FILL if row_offset % 2 == 1 else None

            # Student Info
            cell_name = ws.cell(row=row_idx, column=1, value=student.get("name", ""))
            _apply_cell_style(cell_name, DATA_BOLD_FONT, LEFT, CELL_BORDER, row_fill)
            
            cell_roll = ws.cell(row=row_idx, column=2, value=roll)
            _apply_cell_style(cell_roll, DATA_FONT, CENTER, CELL_BORDER, row_fill)
            
            cell_sec = ws.cell(row=row_idx, column=3, value=student.get("section", ""))
            _apply_cell_style(cell_sec, DATA_FONT, CENTER, CELL_BORDER, row_fill)

            # Date cells
            m_presents = 0
            m_absents = 0
            dates_dict = student.get("dates", {})

            for d_idx, dc in enumerate(m_dates):
                col = 4 + d_idx
                val = dates_dict.get(dc, "-")
                cell = ws.cell(row=row_idx, column=col)
                cell.alignment = CENTER

                if str(val).startswith("P"):
                    if " (" in str(val):
                        time_part = str(val).split(" (")[1].replace(")", "")
                        cell.value = f"P\n{time_part}"
                        cell.font = Font(name="Segoe UI", size=8, bold=True, color="065F46")
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    else:
                        cell.value = "P"
                        cell.font = PRESENT_FONT
                    cell.fill = PRESENT_FILL
                    m_presents += 1
                elif str(val).startswith("A"):
                    cell.value = "A"
                    cell.font = ABSENT_FONT
                    cell.fill = ABSENT_FILL
                    m_absents += 1
                else:
                    cell.value = "–"
                    cell.font = DASH_FONT
                    cell.fill = DASH_FILL
                cell.border = CELL_BORDER

            m_total = m_presents + m_absents
            m_pct = round((m_presents / m_total) * 100, 1) if m_total > 0 else 0.0

            # Month Stats
            c_idx = 4 + num_date_cols
            cell_mp = ws.cell(row=row_idx, column=c_idx, value=m_presents)
            _apply_cell_style(cell_mp, Font(name="Segoe UI", size=10, bold=True, color="065F46"), CENTER, CELL_BORDER, PRESENTS_FILL)

            c_idx += 1
            cell_ma = ws.cell(row=row_idx, column=c_idx, value=m_absents)
            _apply_cell_style(cell_ma, Font(name="Segoe UI", size=10, bold=True, color="991B1B"), CENTER, CELL_BORDER, ABSENTS_FILL)

            c_idx += 1
            cell_mpct = ws.cell(row=row_idx, column=c_idx, value=f"{m_pct}%")
            _apply_cell_style(cell_mpct, PCT_GREEN_FONT if m_pct >= 75 else (PCT_YELLOW_FONT if m_pct >= 50 else PCT_RED_FONT), CENTER, CELL_BORDER, PCT_GREEN_FILL if m_pct >= 75 else (PCT_YELLOW_FILL if m_pct >= 50 else PCT_RED_FILL))



        # COLUMN WIDTHS
        ws.column_dimensions[get_column_letter(1)].width = 22
        ws.column_dimensions[get_column_letter(2)].width = 14
        ws.column_dimensions[get_column_letter(3)].width = 18
        for d_idx in range(num_date_cols):
            ws.column_dimensions[get_column_letter(4 + d_idx)].width = 8
            
        c_idx = 4 + num_date_cols
        ws.column_dimensions[get_column_letter(c_idx)].width = 16
        ws.column_dimensions[get_column_letter(c_idx+1)].width = 16
        ws.column_dimensions[get_column_letter(c_idx+2)].width = 14

    return wb


# =========================
# 📦 READ EXISTING DATA FROM WORKBOOK
# =========================
def _read_existing_data(file_path: str) -> dict:
    """
    Returns {roll: {name, section, dates: {date_str: 'P'/'A'/'-'}} }
    Iterates through all monthly sheets to reconstruct full history.
    """
    students_data = {}

    try:
        wb = load_workbook(file_path)
        
        for sheet_name in wb.sheetnames:
            if sheet_name == "Summary":
                continue
                
            ws = wb[sheet_name]

            # Find the header row
            header_row_idx = None
            for r in ws.iter_rows(min_row=1, max_row=15):
                if r[0].value == "Student Name":
                    header_row_idx = r[0].row
                    break
            if header_row_idx is None:
                continue

            headers_row = [cell.value for cell in ws[header_row_idx]]

            date_headers = []
            for i in range(3, len(headers_row)):
                val = headers_row[i]
                if val and val not in (
                    "Total Presents (in this month)", "Total Absents (in this month)", "Percentage (in this month)",
                    "Total Present (throughout Session)", "Total Absent (throughout Session)", "Percentage (throughout Session)"
                ):
                    date_headers.append((i, val))
                elif val == "Total Presents (in this month)":
                    break

            def _parse_header_to_date(header_text: str, sheet_year: int) -> str:
                try:
                    dt = datetime.strptime(f"{header_text} {sheet_year}", "%d %b %Y")
                    return dt.strftime("%Y-%m-%d")
                except ValueError:
                    return str(header_text)
                    
            sheet_year = datetime.now().year
            try:
                sheet_year = int(sheet_name.split()[-1])
            except:
                pass

            date_col_map = [(col_idx, _parse_header_to_date(ht, sheet_year)) for col_idx, ht in date_headers]

            data_start = header_row_idx + 1
            for row in ws.iter_rows(min_row=data_start, values_only=False):
                name_val = row[0].value
                roll_val = row[1].value
                section_val = row[2].value
                if not roll_val:
                    continue
                    
                roll_str = str(roll_val).strip()
                if roll_str not in students_data:
                    students_data[roll_str] = {
                        "name": name_val or "",
                        "section": section_val or "",
                        "dates": {}
                    }
                    
                for col_idx, date_str in date_col_map:
                    cell_val = row[col_idx].value
                    if cell_val:
                        students_data[roll_str]["dates"][date_str] = str(cell_val).strip()

    except Exception as e:
        print(f"Warning reading workbook: {e}")

    return students_data


# =========================
# 📄 GET FILE PATH
# =========================
def _get_file_path(
    class_name: str, 
    department: str, 
    section: str = "N/A", 
    academic_session: str = "N/A", 
    course_code: str = "N/A"
) -> str:
    os.makedirs(CSV_DIR, exist_ok=True)
    
    class_name = class_name or "unknown_class"
    department = department or "Dept"
    section = section if section and section != "N/A" else ""
    academic_session = academic_session if academic_session and academic_session != "N/A" else ""
    course_code = course_code if course_code and course_code != "N/A" else ""
    
    dept_sec = f"{department}_{section}" if section else department
    session_part = academic_session
    class_part = f"{class_name} ({course_code})" if course_code else class_name
    
    parts = [p for p in [dept_sec, session_part, class_part] if p]
    filename = ",".join(parts) + "_Register.xlsx"
    
    safe_filename = _safe_filename(filename)
    return os.path.join(CSV_DIR, safe_filename)


# =========================
# 📝 INITIALIZE CLASS REGISTER
# =========================
def initialize_class_csv(
    *,
    class_name: str,
    students: list,
    date: str,
    section: str,
    department: str,
    course_code: str = "N/A",
    semester: str = "N/A",
    year: str = "N/A",
    academic_session: str = "N/A",
):
    file_path = _get_file_path(class_name, department, section, academic_session, course_code)
    date_columns = _generate_date_columns(date)

    existing_data = {}
    if os.path.isfile(file_path):
        existing_data = _read_existing_data(file_path)

    for student in students:
        roll = str(student.get("roll", "")).strip()
        if not roll:
            continue
        if roll not in existing_data:
            existing_data[roll] = {
                "name": student.get("name", "Unknown"),
                "section": section or "N/A",
                "dates": {}
            }
        today_val = existing_data[roll]["dates"].get(date, "-")
        if today_val in ("-", "", None):
            existing_data[roll]["dates"][date] = "A"

    all_dates_set = set(date_columns)
    for rd in existing_data.values():
        all_dates_set.update(rd.get("dates", {}).keys())
    date_columns = sorted(all_dates_set)

    try:
        wb = _build_workbook(
            class_name=class_name,
            department=department,
            section=section,
            date_columns=date_columns,
            students_data=existing_data,
            course_code=course_code,
            semester=semester,
            year=year,
            academic_session=academic_session,
        )
        wb.save(file_path)
        print(f"Excel Register Initialized: {file_path}")
    except Exception as e:
        print(f"Error writing Excel register: {e}")


# =========================
# 📝 MARK PRESENT
# =========================
def append_attendance_sheet(
    *,
    student_id: str,
    name: str,
    roll: str,
    class_name: str,
    date: str,
    time: str,
    section: str,
    department: str,
    course_code: str = "N/A",
    semester: str = "N/A",
    year: str = "N/A",
    academic_session: str = "N/A",
):
    file_path = _get_file_path(class_name, department, section, academic_session, course_code)
    date_columns = _generate_date_columns(date)

    existing_data = {}
    if os.path.isfile(file_path):
        existing_data = _read_existing_data(file_path)

    roll = str(roll).strip()
    if roll not in existing_data:
        existing_data[roll] = {
            "name": name,
            "section": section or "N/A",
            "dates": {}
        }

    if time and time != "N/A":
        existing_data[roll]["dates"][date] = f"P ({time})"
    else:
        existing_data[roll]["dates"][date] = "P"

    all_dates_set = set(date_columns)
    for rd in existing_data.values():
        all_dates_set.update(rd.get("dates", {}).keys())
    date_columns = sorted(all_dates_set)

    try:
        wb = _build_workbook(
            class_name=class_name,
            department=department,
            section=section,
            date_columns=date_columns,
            students_data=existing_data,
            course_code=course_code,
            semester=semester,
            year=year,
            academic_session=academic_session,
        )
        wb.save(file_path)
        print(f"Excel Register Updated (Present): {file_path}")
    except Exception as e:
        print(f"Error writing Excel register: {e}")


# =========================
# 🔄 SYNC MANUAL OVERRIDE
# =========================
def sync_csv_manual_override(
    *,
    class_name: str,
    roll: str,
    date: str,
    new_status: str,
    time: str = "N/A",
    department: str = "",
    section: str = "",
    course_code: str = "N/A",
    semester: str = "N/A",
    year: str = "N/A",
    academic_session: str = "N/A",
):
    file_path = _get_file_path(class_name, department, section, academic_session, course_code)
    if not os.path.isfile(file_path):
        return

    date_columns = _generate_date_columns(date)
    existing_data = _read_existing_data(file_path)

    roll = str(roll).strip()
    if roll not in existing_data:
        return

    all_dates_set = set(date_columns)
    for rd in existing_data.values():
        all_dates_set.update(rd.get("dates", {}).keys())
    date_columns = sorted(all_dates_set)

    if new_status == "Present":
        existing_data[roll]["dates"][date] = f"P ({time})" if time and time != "N/A" else "P"
    else:
        existing_data[roll]["dates"][date] = "A"

    used_section = section or existing_data[roll].get("section", "")

    try:
        wb = _build_workbook(
            class_name=class_name,
            department=department,
            section=used_section,
            date_columns=date_columns,
            students_data=existing_data,
            course_code=course_code,
            semester=semester,
            year=year,
            academic_session=academic_session,
        )
        wb.save(file_path)
        print(f"Excel Register Overridden ({new_status}): {file_path}")
    except Exception as e:
        print(f"Error writing Excel override: {e}")
